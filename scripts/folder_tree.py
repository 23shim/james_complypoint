"""Generate deduplicated folder tree Excel files at 2 and 3 levels deep.

Each folder is annotated with entity types (scheme, plot, address).
Scheme, address, and plot names cascade down to child rows so you can
scan the column and see which scheme/address context each row belongs to.
"""

import json
import sys
import os
from collections import defaultdict
from typing import Any

sys.stdout.reconfigure(encoding="utf-8")

import pandas as pd

SEP = "\\"

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Entity lookup builders ────────────────────────────────────────────────────

def _build_entity_lookups(df):
    """Build folder_path -> entity info lookups from classified data.

    Returns three dicts (scheme, plot, address), each mapping
    a relative folder path to a dict with keys: values, confidences, count.
    """
    scheme_info  = defaultdict(lambda: {"values": set(), "confidences": [], "count": 0})
    plot_info    = defaultdict(lambda: {"values": set(), "confidences": [], "count": 0})
    address_info = defaultdict(lambda: {"values": set(), "confidences": [], "count": 0})

    for _, row in df.iterrows():
        segs = row.get("segments")
        if not isinstance(segs, list) or not segs:
            continue

        # Scheme: use entity_scheme_path directly (already relative)
        scheme_path = row.get("entity_scheme_path", "")
        if scheme_path:
            scheme_info[scheme_path]["values"].add(row.get("entity_scheme", ""))
            scheme_info[scheme_path]["confidences"].append(
                row.get("entity_scheme_confidence", 0.0)
            )
            scheme_info[scheme_path]["count"] += 1

        # Plot: reconstruct folder path from segments[:depth+1]
        plot_val   = row.get("entity_plot", "")
        plot_depth = row.get("entity_plot_depth", -1)
        if plot_val and isinstance(plot_depth, (int, float)) and int(plot_depth) >= 0:
            d = int(plot_depth)
            if d < len(segs):
                plot_path = SEP.join(segs[: d + 1])
                plot_info[plot_path]["values"].add(plot_val)
                plot_info[plot_path]["confidences"].append(
                    row.get("entity_plot_confidence", 0.0)
                )
                plot_info[plot_path]["count"] += 1

        # Address: reconstruct folder path from segments[:depth+1]
        addr_val   = row.get("entity_address", "")
        addr_depth = row.get("entity_address_depth", -1)
        if addr_val and isinstance(addr_depth, (int, float)) and int(addr_depth) >= 0:
            d = int(addr_depth)
            if d < len(segs):
                addr_path = SEP.join(segs[: d + 1])
                address_info[addr_path]["values"].add(addr_val)
                address_info[addr_path]["confidences"].append(
                    row.get("entity_address_confidence", 0.0)
                )
                address_info[addr_path]["count"] += 1

    return scheme_info, plot_info, address_info


def _build_raw_detection_lookups(df):
    """Build folder_path -> raw detection info from raw_plots/raw_addresss.

    These capture EVERY address/plot match at every depth — including
    candidates that were not selected as the organisation-level entity
    (suppressed by validation, hierarchy enforcement, or a deeper match
    winning).

    Returns two dicts (raw_plot_info, raw_address_info), each mapping
    a relative folder path to {values, confidences, count}.
    """
    raw_plot_info    = defaultdict(lambda: {"values": set(), "confidences": [], "count": 0})
    raw_address_info = defaultdict(lambda: {"values": set(), "confidences": [], "count": 0})

    has_raw_plots = "raw_plots" in df.columns
    has_raw_addrs = "raw_addresss" in df.columns

    if not has_raw_plots and not has_raw_addrs:
        return raw_plot_info, raw_address_info

    for _, row in df.iterrows():
        segs = row.get("segments")
        if not isinstance(segs, list) or not segs:
            continue

        # Raw plot detections
        if has_raw_plots:
            for match in (row.get("raw_plots") or []):
                if not isinstance(match, dict):
                    continue
                d = match.get("depth", -1)
                if isinstance(d, (int, float)) and int(d) >= 0 and int(d) < len(segs):
                    path = SEP.join(segs[: int(d) + 1])
                    raw_plot_info[path]["values"].add(match.get("value", ""))
                    raw_plot_info[path]["confidences"].append(
                        match.get("confidence", 0.0)
                    )
                    raw_plot_info[path]["count"] += 1

        # Raw address detections
        if has_raw_addrs:
            for match in (row.get("raw_addresss") or []):
                if not isinstance(match, dict):
                    continue
                d = match.get("depth", -1)
                if isinstance(d, (int, float)) and int(d) >= 0 and int(d) < len(segs):
                    path = SEP.join(segs[: int(d) + 1])
                    raw_address_info[path]["values"].add(match.get("value", ""))
                    raw_address_info[path]["confidences"].append(
                        match.get("confidence", 0.0)
                    )
                    raw_address_info[path]["count"] += 1

    return raw_plot_info, raw_address_info


def _summarise(info_dict, rel_path):
    """Exact-match only: return (value, confidence, count) or ('', 0, 0)."""
    if rel_path not in info_dict:
        return "", 0.0, 0
    entry = info_dict[rel_path]
    value = "; ".join(sorted(v for v in entry["values"] if v))
    conf  = round(sum(entry["confidences"]) / len(entry["confidences"]), 2) \
            if entry["confidences"] else 0.0
    return value, conf, entry["count"]


def _summarise_cascade(info_dict, rel_path):
    """Return (value, confidence, count) for rel_path or its nearest ancestor."""
    # Try exact match first
    if rel_path in info_dict:
        return _summarise(info_dict, rel_path)
    # Walk up the tree
    parts = rel_path.split(SEP)
    for depth in range(len(parts) - 1, 0, -1):
        ancestor = SEP.join(parts[:depth])
        if ancestor in info_dict:
            return _summarise(info_dict, ancestor)
    return "", 0.0, 0


# ── Scheme candidates sidecar ─────────────────────────────────────────────────

def _load_scheme_candidates(path: str | None) -> dict[str, dict[str, Any]]:
    """Load the scheme-candidates JSON sidecar written by the pipeline.

    Returns a dict mapping relative folder path → candidate fields dict.
    Returns an empty dict if no path is given or the file cannot be read.
    """
    if not path:
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        print(f"  Warning: could not load scheme candidates ({exc})")
        return {}


# Scoring constants duplicated here so folder_tree.py has no src/ dependency.
# !! Keep in sync with scheme_detector.py !!
_DIRECT_PLOT_CHILDREN_BOOST    = 0.10
_DEVELOPMENT_CATEGORIES_BOOST  = 0.10
_HIGH_ENTITY_CONFIDENCE        = 0.85
_MED_ENTITY_CONFIDENCE         = 0.70


# Reason codes written by _identify_candidates for folders that passed the
# structural gate but were excluded before scoring.  Mapped to human-readable
# descriptions shown in the score trace column.
_PRE_FILTER_REASONS: dict[str, str] = {
    "too_short":   "filtered: folder name too short",
    "cfg_excl":    "filtered: config exclusion (dates, containers, status labels)",
    "cat_folder":  "filtered: matches category/type signal",
    "nested":      "filtered: nested inside category/type ancestor",
    "entity_only": "filtered: pure plot entity (not a scheme name)",
}


def _build_score_trace(m: dict[str, Any]) -> str:
    """Build a compact scoring trace from a scheme-candidate dict.

    Mirrors scheme_detector.build_score_trace() but works from the
    plain dict loaded from the JSON sidecar (no src/ import needed).

    For pre-filter excluded entries (passed structural gate but not scored)
    returns a short human-readable reason instead of a scoring breakdown.
    """
    status = m.get("candidate_status", "")
    if status in _PRE_FILTER_REASONS:
        reason_text = _PRE_FILTER_REASONS[status]
        unique_ents = m.get("unique_entities", 0)
        return f"{reason_text}; {unique_ents} entities"

    parts: list[str] = []

    conf = m.get("confidence", 0.0)
    if conf >= _HIGH_ENTITY_CONFIDENCE:
        parts.append(f"base={_HIGH_ENTITY_CONFIDENCE:.2f}")
    else:
        parts.append(f"base={_MED_ENTITY_CONFIDENCE:.2f}")

    dir_plots = m.get("direct_plot_children", 0)
    if dir_plots > 0:
        parts.append(f"dir_plots({dir_plots})=+{_DIRECT_PLOT_CHILDREN_BOOST:.2f}")

    if m.get("place_name_match"):
        parts.append("place_name=yes")

    if m.get("address_like_name"):
        parts.append("addr_match=yes")

    return "; ".join(parts) + f" → {conf:.3f}"


# ── Excel writer ──────────────────────────────────────────────────────────────

def _write_xlsx(out_df, out_path, max_depth):
    """Write a formatted Excel file. Falls back to CSV if openpyxl unavailable."""
    if not HAS_OPENPYXL:
        csv_path = out_path.replace(".xlsx", ".csv")
        out_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"  (openpyxl not available — wrote CSV: {csv_path})")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Folder Tree ({max_depth} levels)"

    # ── Colour palette ────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")  # dark navy
    LEVEL_FILL   = PatternFill("solid", fgColor="D6E4F0")  # light blue — level columns
    SCHEME_FILL  = PatternFill("solid", fgColor="FFF2CC")  # pale yellow — scheme rows
    ADDR_FILL    = PatternFill("solid", fgColor="E2EFDA")  # pale green  — address rows
    PLOT_FILL    = PatternFill("solid", fgColor="FCE4D6")  # pale orange — plot rows
    ALT_FILL     = PatternFill("solid", fgColor="F5F5F5")  # light grey  — alternating

    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    BODY_FONT    = Font(size=10, name="Calibri")

    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")

    thin        = Side(style="thin", color="D0D0D0")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Column widths ─────────────────────────────────────────────────────
    WIDTH = {
        "folder_path":                65,
        "depth":                       5,
        "entity_types":               20,
        "is_scheme":                   9,
        "scheme_name":                38,
        "scheme_confidence":           8,
        "scheme_file_count":           8,
        "scheme_candidate_score":     10,
        "scheme_candidate_status":    12,
        "scheme_score_trace":         80,
        "detected_plot":              22,
        "detected_plot_confidence":    8,
        "is_plot":                     7,
        "plot_name":                  22,
        "plot_confidence":             8,
        "plot_file_count":             8,
        "detected_address":           38,
        "detected_address_confidence": 8,
        "is_address":                 10,
        "address_name":               38,
        "address_confidence":          8,
        "address_file_count":          8,
    }
    for i in range(1, max_depth + 1):
        WIDTH[f"level_{i}"] = 32

    cols = list(out_df.columns)

    # ── Write header ──────────────────────────────────────────────────────
    ws.append(cols)
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = WIDTH.get(col_name, 14)

    # ── Identify key column indices (1-based) ─────────────────────────────
    def _idx(name):
        return cols.index(name) + 1 if name in cols else None

    level_col_indices = {_idx(f"level_{i}") for i in range(1, max_depth + 1)} - {None}
    is_scheme_idx      = _idx("is_scheme")
    is_addr_idx        = _idx("is_address")
    is_plot_idx        = _idx("is_plot")
    det_addr_idx       = _idx("detected_address")
    det_plot_idx       = _idx("detected_plot")
    conf_cols      = {"scheme_confidence", "address_confidence", "plot_confidence",
                      "detected_address_confidence", "detected_plot_confidence",
                      "scheme_candidate_score"}
    count_cols     = {"scheme_file_count", "address_file_count", "plot_file_count",
                      "depth"}
    flag_cols      = {"is_scheme", "is_address", "is_plot", "scheme_candidate_status"}

    # ── Write data rows ───────────────────────────────────────────────────
    for ri, row_data in enumerate(out_df.itertuples(index=False), start=2):
        ws.append(list(row_data))

        is_scheme_row = is_scheme_idx and ws.cell(ri, is_scheme_idx).value == "Yes"
        is_addr_row   = is_addr_idx   and ws.cell(ri, is_addr_idx).value   == "Yes"
        is_plot_row   = is_plot_idx   and ws.cell(ri, is_plot_idx).value   == "Yes"
        has_det_addr  = det_addr_idx  and bool(ws.cell(ri, det_addr_idx).value)
        has_det_plot  = det_plot_idx  and bool(ws.cell(ri, det_plot_idx).value)

        if is_scheme_row:
            row_fill = SCHEME_FILL
        elif is_addr_row or has_det_addr:
            row_fill = ADDR_FILL
        elif is_plot_row or has_det_plot:
            row_fill = PLOT_FILL
        elif ri % 2 == 0:
            row_fill = ALT_FILL
        else:
            row_fill = None

        ws.row_dimensions[ri].height = 15

        for ci, col_name in enumerate(cols, 1):
            cell = ws.cell(ri, ci)
            cell.font   = BODY_FONT
            cell.border = thin_border

            if ci in level_col_indices:
                cell.alignment = LEFT
                cell.fill = LEVEL_FILL if row_fill is None else row_fill
            elif col_name in flag_cols or col_name in count_cols:
                cell.alignment = CENTER
                if row_fill:
                    cell.fill = row_fill
            elif col_name in conf_cols:
                cell.alignment = CENTER
                if isinstance(cell.value, (int, float)) and cell.value:
                    cell.number_format = "0.00"
                if row_fill:
                    cell.fill = row_fill
            else:
                cell.alignment = LEFT
                if row_fill:
                    cell.fill = row_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(out_path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    os.chdir(os.path.join(os.path.dirname(__file__), ".."))

    if len(sys.argv) < 2:
        sys.exit("Usage: python scripts/folder_tree.py <classified.jsonl> [root_prefix] [--depth N] [--scheme-candidates path]")
    input_path  = sys.argv[1]
    root_prefix = sys.argv[2] if len(sys.argv) > 2 else ""
    # Optional: --depth 3  to generate only the 3-levels file
    only_depth = None
    if "--depth" in sys.argv:
        idx = sys.argv.index("--depth")
        if idx + 1 < len(sys.argv):
            only_depth = int(sys.argv[idx + 1])
    # Optional: --scheme-candidates path  to add confidence/trace columns
    candidates_path = None
    if "--scheme-candidates" in sys.argv:
        idx = sys.argv.index("--scheme-candidates")
        if idx + 1 < len(sys.argv):
            candidates_path = sys.argv[idx + 1]

    stem = os.path.splitext(os.path.basename(input_path))[0].replace("_classified", "")

    # Only load the columns needed for entity lookup — avoids OOM on large JSONL
    KEEP = {
        "segments",
        "entity_scheme_path", "entity_scheme", "entity_scheme_confidence",
        "entity_plot", "entity_plot_depth", "entity_plot_confidence",
        "entity_address", "entity_address_depth", "entity_address_confidence",
        "raw_plots", "raw_addresss",
    }
    records = []
    with open(input_path, encoding="utf-8") as f:
        for line in f:
            if line.strip():
                full = json.loads(line)
                records.append({k: full[k] for k in KEEP if k in full})
    df = pd.DataFrame(records)
    print(f"Loaded {len(df):,} records from {input_path}")

    print("Building entity lookups...")
    scheme_info, plot_info, address_info = _build_entity_lookups(df)
    raw_plot_info, raw_address_info = _build_raw_detection_lookups(df)
    print(
        f"  Scheme folders: {len(scheme_info):,}, "
        f"Plot folders: {len(plot_info):,}, "
        f"Address folders: {len(address_info):,}"
    )
    if raw_plot_info or raw_address_info:
        print(
            f"  Raw detections: {len(raw_plot_info):,} plot folders, "
            f"{len(raw_address_info):,} address folders"
        )

    scheme_candidates = _load_scheme_candidates(candidates_path)
    if scheme_candidates:
        print(f"  Scheme candidates loaded: {len(scheme_candidates):,} entries")

    ROOT       = root_prefix if root_prefix else ""
    root_depth = ROOT.count(SEP) if ROOT else 0

    for max_depth in ((only_depth,) if only_depth else (2, 3, 4)):
        folders: set = set()
        for segs in df["segments"]:
            if not isinstance(segs, list) or not segs:
                continue
            for level in range(1, min(max_depth + 1, len(segs) + 1)):
                folder_path = (ROOT + SEP + SEP.join(segs[:level])) if ROOT \
                              else SEP.join(segs[:level])
                folders.add(folder_path)

        folder_list = sorted(folders)
        out_df = pd.DataFrame({"folder_path": folder_list})

        out_df["depth"] = out_df["folder_path"].apply(
            lambda p: p.count(SEP) - root_depth
        )

        for i in range(1, max_depth + 1):
            out_df[f"level_{i}"] = out_df["folder_path"].apply(
                lambda p, lvl=i: (
                    p.split(SEP)[root_depth + lvl]
                    if len(p.split(SEP)) > root_depth + lvl else ""
                )
            )

        # Relative path for entity lookups
        if ROOT:
            prefix   = ROOT + SEP
            rel_paths = out_df["folder_path"].apply(
                lambda p: p[len(prefix):] if p.startswith(prefix) else p
            )
        else:
            rel_paths = out_df["folder_path"]

        # ── Scheme columns ────────────────────────────────────────────────
        # is_scheme / confidence / file_count: exact match only
        # scheme_name: cascade from nearest ancestor
        scheme_exact    = rel_paths.apply(lambda rp: _summarise(scheme_info, rp))
        scheme_cascaded = rel_paths.apply(lambda rp: _summarise_cascade(scheme_info, rp))

        out_df["is_scheme"]          = scheme_exact.apply(lambda x: "Yes" if x[0] else "")
        out_df["scheme_name"]        = scheme_cascaded.apply(lambda x: x[0])
        out_df["scheme_confidence"]  = scheme_exact.apply(lambda x: x[1] if x[0] else "")
        out_df["scheme_file_count"]  = scheme_exact.apply(lambda x: x[2] if x[0] else "")

        # Candidate score + reasoning trace for THIS folder (exact match only).
        # scheme_candidate_score: confidence the algorithm gave this folder as a
        #   scheme candidate (present for all surviving candidates, not just winners).
        # scheme_score_trace: step-by-step breakdown showing every signal's
        #   contribution so mis-scores can be diagnosed.
        if scheme_candidates:
            def _candidate_score(rp: str):
                cand = scheme_candidates.get(rp)
                if not cand:
                    return ""
                # Pre-filter entries have confidence=0.0 but were never scored
                if cand.get("candidate_status", "") in _PRE_FILTER_REASONS:
                    return ""
                return round(cand["confidence"], 3)

            def _candidate_status(rp: str) -> str:
                cand = scheme_candidates.get(rp)
                return cand.get("candidate_status", "") if cand else ""

            def _candidate_trace(rp: str) -> str:
                cand = scheme_candidates.get(rp)
                return _build_score_trace(cand) if cand else ""

            out_df["scheme_candidate_score"]  = rel_paths.apply(_candidate_score)
            out_df["scheme_candidate_status"] = rel_paths.apply(_candidate_status)
            out_df["scheme_score_trace"]      = rel_paths.apply(_candidate_trace)
        else:
            out_df["scheme_candidate_score"]  = ""
            out_df["scheme_candidate_status"] = ""
            out_df["scheme_score_trace"]      = ""

        # ── Plot columns ──────────────────────────────────────────────────
        # detected_plot: raw match at THIS folder (before validation)
        # plot_name:     org-level assignment (cascaded, after validation)
        raw_plot_exact = rel_paths.apply(lambda rp: _summarise(raw_plot_info, rp))
        out_df["detected_plot"]            = raw_plot_exact.apply(lambda x: x[0])
        out_df["detected_plot_confidence"] = raw_plot_exact.apply(
            lambda x: x[1] if x[0] else ""
        )

        plot_exact    = rel_paths.apply(lambda rp: _summarise(plot_info, rp))
        plot_cascaded = rel_paths.apply(lambda rp: _summarise_cascade(plot_info, rp))

        out_df["is_plot"]          = plot_exact.apply(lambda x: "Yes" if x[0] else "")
        out_df["plot_name"]        = plot_cascaded.apply(lambda x: x[0])
        out_df["plot_confidence"]  = plot_exact.apply(lambda x: x[1] if x[0] else "")
        out_df["plot_file_count"]  = plot_exact.apply(lambda x: x[2] if x[0] else "")

        # ── Address columns ───────────────────────────────────────────────
        # detected_address: raw match at THIS folder (before validation)
        # address_name:     org-level assignment (cascaded, after validation)
        raw_addr_exact = rel_paths.apply(lambda rp: _summarise(raw_address_info, rp))
        out_df["detected_address"]            = raw_addr_exact.apply(lambda x: x[0])
        out_df["detected_address_confidence"] = raw_addr_exact.apply(
            lambda x: x[1] if x[0] else ""
        )

        addr_exact    = rel_paths.apply(lambda rp: _summarise(address_info, rp))
        addr_cascaded = rel_paths.apply(lambda rp: _summarise_cascade(address_info, rp))

        out_df["is_address"]          = addr_exact.apply(lambda x: "Yes" if x[0] else "")
        out_df["address_name"]        = addr_cascaded.apply(lambda x: x[0])
        out_df["address_confidence"]  = addr_exact.apply(lambda x: x[1] if x[0] else "")
        out_df["address_file_count"]  = addr_exact.apply(lambda x: x[2] if x[0] else "")

        # ── entity_types summary ──────────────────────────────────────────
        def _entity_types(row):
            types = []
            if row["is_scheme"]:      types.append("scheme")
            if row["is_plot"]:        types.append("plot")
            elif row["detected_plot"]:types.append("detected_plot")
            if row["is_address"]:        types.append("address")
            elif row["detected_address"]:types.append("detected_address")
            return ", ".join(types)

        out_df["entity_types"] = out_df.apply(_entity_types, axis=1)

        # ── Write XLSX ────────────────────────────────────────────────────
        os.makedirs("reports", exist_ok=True)
        out_path = f"reports/folder_tree_{stem}_{max_depth}_levels.xlsx"
        _write_xlsx(out_df, out_path, max_depth)

        with_entity = (out_df["entity_types"] != "").sum()
        print(
            f"Wrote {len(out_df):,} unique folders ({max_depth} levels) to {out_path}"
            f"  — {with_entity:,} have entity labels"
        )


if __name__ == "__main__":
    main()

"""
leaf_folders.py — Extract the deepest (leaf) folder paths from a TreeSize XLSX
or a raw path file.

A "leaf folder" is one that has no sub-folders in the dataset — the deepest
point down every branch of the directory tree.  Intermediate parent folders are
dropped, so the output is the minimal set of folder paths that covers the entire
estate.

Modes
-----
  With --config (XLSX mode):
      Reads the XLSX using column mapping and header-row offset from schema.yaml,
      and strips the root_prefix defined in run_config.yaml from the output.

  Without --config (raw mode):
      Accepts a plain text file (one path per line) or a CSV whose first column
      is named "Path" or "path".  No prefix stripping is applied unless
      --root-prefix is passed explicitly.

Usage
-----
  # XLSX mode (recommended — respects schema + strips prefix)
  python scripts/leaf_folders.py \\
      --config config/run_config.yaml \\
      --input  context/wkh_apps1_wkh_apps1Techserv_Development.xlsx \\
      --output output/leaf_folders.csv

  # Raw mode
  python scripts/leaf_folders.py \\
      --input paths.txt \\
      --output output/leaf_folders.csv
"""

import argparse
import csv
import os
import sys
from pathlib import Path


# ── Path helpers ─────────────────────────────────────────────────────────────

def is_valid_path(value) -> bool:
    """Return True if value looks like a real filesystem path (not a corrupted row)."""
    if not isinstance(value, str):
        return False
    return "\\" in value or "/" in value


def folder_of(path: str) -> str:
    """
    Return the folder component of a path.
    If the last segment has an extension we treat it as a file and take its
    parent.  Otherwise we treat the whole path as a folder.
    """
    p = Path(path.replace("/", "\\"))
    if p.suffix:          # has extension → file, take parent
        return str(p.parent)
    return str(p)


def strip_root(path: str, prefix: str) -> str:
    """
    Strip root_prefix from path (case-insensitive).
    Returns a relative path, or the original if the prefix is not found.
    """
    if not prefix:
        return path
    norm_path   = path.replace("/", "\\").rstrip("\\")
    norm_prefix = prefix.replace("/", "\\").rstrip("\\")
    if norm_path.lower().startswith(norm_prefix.lower()):
        relative = norm_path[len(norm_prefix):].lstrip("\\")
        return relative or "\\"
    return path


# ── Leaf-folder algorithm ─────────────────────────────────────────────────────

def find_leaf_folders(folders: set) -> list:
    """
    Return only the leaf folders — those with no sub-folders in the set.

    Algorithm (O(n log n)):
      Normalise paths to lowercase + backslash, then sort.  Because `\\` (ASCII
      92) is lower than every lowercase letter (97-122), all descendants of a
      folder sort immediately after it, before any sibling whose name starts
      with a later letter.  So we only need to check the *next* item to decide
      whether the current folder has children.
    """
    sep = "\\"
    # Build (normalised_key, original_path) pairs and sort by key.
    pairs = sorted(
        ((f.replace("/", "\\").rstrip("\\").lower(), f) for f in folders),
        key=lambda x: x[0],
    )
    leaves = []
    n = len(pairs)
    for i, (norm, original) in enumerate(pairs):
        child_prefix = norm + sep
        if i + 1 < n and pairs[i + 1][0].startswith(child_prefix):
            continue          # has at least one child in the set — not a leaf
        leaves.append(original)
    return leaves


# ── Readers ───────────────────────────────────────────────────────────────────

def read_xlsx(input_path: str, header_row: int, sheet_strategy: str,
              path_column: str) -> list:
    """Read paths from a TreeSize XLSX export."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required for XLSX input: pip install openpyxl")

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    if sheet_strategy == "largest":
        sheet = max(wb.worksheets, key=lambda ws: ws.max_row)
    else:
        sheet = wb[sheet_strategy]

    row_iter = sheet.iter_rows(values_only=True)
    for _ in range(header_row):           # skip metadata rows before headers
        next(row_iter, None)

    raw_headers = next(row_iter, [])
    headers = [str(c).strip() if c is not None else "" for c in raw_headers]

    try:
        col_idx = headers.index(path_column)
    except ValueError:
        sys.exit(
            f"Column '{path_column}' not found in sheet.  "
            f"Available headers: {headers}"
        )

    paths = []
    for row in row_iter:
        val = row[col_idx] if col_idx < len(row) else None
        if val is not None:
            paths.append(str(val))

    wb.close()
    return paths


def read_raw(input_path: str) -> list:
    """
    Read paths from a plain text file (one per line) or a CSV whose first
    column is named Path / path / full_path.
    """
    paths = []
    with open(input_path, "r", encoding="utf-8-sig", errors="replace") as fh:
        first = fh.readline()
        fh.seek(0)
        lower_first = first.strip().lower()
        if "," in first and any(
            lower_first.startswith(k) for k in ("path", "full_path")
        ):
            reader = csv.DictReader(fh)
            for row in reader:
                val = (
                    row.get("Path")
                    or row.get("path")
                    or row.get("full_path")
                    or row.get("Full Path")
                )
                if val:
                    paths.append(val.strip())
        else:
            for line in fh:
                line = line.strip()
                if line:
                    paths.append(line)
    return paths


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Extract leaf folder paths from a TreeSize XLSX or raw path file."
    )
    parser.add_argument(
        "--input", required=True,
        help="Input file (XLSX, CSV, or plain text with one path per line)",
    )
    parser.add_argument(
        "--config",
        help=(
            "Path to run_config.yaml.  Enables XLSX column mapping from "
            "schema.yaml and strips root_prefix from output paths."
        ),
    )
    parser.add_argument(
        "--output",
        default=None,
        help=(
            "Output CSV path.  When --config is provided this defaults to "
            "{output.directory}/leaf_folders_{source_system}.csv."
        ),
    )
    parser.add_argument(
        "--root-prefix",
        default="",
        help=(
            "Override root prefix to strip (ignored when --config is provided, "
            "which reads the prefix from run_config.yaml)."
        ),
    )
    args = parser.parse_args()

    # ── Load config if provided ───────────────────────────────────────────
    root_prefix    = args.root_prefix
    header_row     = 4
    sheet_strategy = "largest"
    path_column    = "Path"
    output_path    = args.output or "output/leaf_folders.csv"

    if args.config:
        try:
            import yaml
        except ImportError:
            sys.exit("PyYAML is required when using --config: pip install pyyaml")

        with open(args.config, "r") as fh:
            run_cfg = yaml.safe_load(fh)

        root_prefix   = run_cfg.get("source", {}).get("root_prefix", "")
        source_system = run_cfg.get("source", {}).get("source_system", "output")
        out_dir       = run_cfg.get("output", {}).get("directory", "output")

        # Derive output path from config unless the user explicitly passed --output
        if args.output is None:
            safe_name = source_system.replace(" ", "_")
            output_path = f"{out_dir}/leaf_folders_{safe_name}.csv"

        schema_path = Path(args.config).parent / "schema.yaml"
        if schema_path.exists():
            with open(schema_path, "r") as fh:
                schema_cfg = yaml.safe_load(fh)
            fmt = run_cfg.get("source", {}).get("format", "treesize")
            fmt_cfg = schema_cfg.get("formats", {}).get(fmt, {})
            header_row     = fmt_cfg.get("header_row", 4)
            sheet_strategy = fmt_cfg.get("sheet_strategy", "largest")
            col_map = fmt_cfg.get("column_map", {})
            # Reverse-lookup: find the source column that maps to full_path
            path_column = next(
                (k for k, v in col_map.items() if v == "full_path"), "Path"
            )

    # ── Read ──────────────────────────────────────────────────────────────
    input_ext = Path(args.input).suffix.lower()
    print(f"Reading: {args.input}")

    if input_ext in (".xlsx", ".xlsm", ".xlsb"):
        raw_paths = read_xlsx(args.input, header_row, sheet_strategy, path_column)
    else:
        raw_paths = read_raw(args.input)

    print(f"  Rows read:       {len(raw_paths):>10,}")

    # ── Filter corrupted rows ─────────────────────────────────────────────
    valid_paths = [p for p in raw_paths if is_valid_path(p)]
    skipped = len(raw_paths) - len(valid_paths)
    if skipped:
        print(f"  Corrupted rows:  {skipped:>10,}  (skipped)")

    # ── Extract folder for each path ──────────────────────────────────────
    folders: set = set()
    for path in valid_paths:
        folder = folder_of(path)
        if is_valid_path(folder):
            folders.add(folder.replace("/", "\\").rstrip("\\"))

    print(f"  Unique folders:  {len(folders):>10,}")

    # ── Find leaf folders ─────────────────────────────────────────────────
    leaves = find_leaf_folders(folders)
    print(f"  Leaf folders:    {len(leaves):>10,}")

    # ── Strip root prefix from output ─────────────────────────────────────
    if root_prefix:
        print(f"  Stripping prefix: {root_prefix}")
        output_paths = [strip_root(p, root_prefix) for p in leaves]
    else:
        output_paths = leaves

    # ── Write CSV ─────────────────────────────────────────────────────────
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    with open(out, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(["folder_path"])
        for path in output_paths:
            writer.writerow([path])

    input_count = len(raw_paths)
    reduction = 100.0 * (1 - len(leaves) / input_count) if input_count else 0
    print(f"\nOutput: {out}")
    print(
        f"  {input_count:,} input rows  ->  {len(leaves):,} leaf folders  "
        f"({reduction:.1f}% reduction)"
    )


if __name__ == "__main__":
    main()
